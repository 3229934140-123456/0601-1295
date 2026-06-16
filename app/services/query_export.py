from __future__ import annotations
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
import uuid
import io
import os
import csv
import json
from collections import defaultdict
from sqlalchemy import select, and_, or_, func, desc, asc, between
from sqlalchemy.ext.asyncio import AsyncSession
from app.core import logger, settings
from app.core.constants import (
    SeverityLevel, TicketStatus, ViolationResult, EventType,
    DisciplinaryAction, LogActionType
)
from app.core.database import get_db_context
from app.models.investigation import (
    InvestigationTicket, ComplianceEvent, EvidenceItem,
    EvidencePackage
)
from app.models.compliance import (
    SystemLog, EmployeeReport, DailyStatistics,
    ComplianceProfile, DisciplinaryActionRecord
)
from app.models.organization import Employee, Department
from app.models.data_source import (
    EmailRecord, InstantMessageRecord,
    DoorAccessRecord, FinanceRecord
)


class QueryAndExportService:
    MAX_EXPORT_ROWS = 100000

    def __init__(self):
        self.logger = logger.bind(module="QueryAndExportService")

    async def query_tickets(
        self,
        employee_id: Optional[uuid.UUID] = None,
        event_type: Optional[str] = None,
        severity: Optional[str] = None,
        status: Optional[str] = None,
        department_id: Optional[uuid.UUID] = None,
        assigned_officer_id: Optional[uuid.UUID] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        date_field: str = "created_at",
        keyword: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> Tuple[List[InvestigationTicket], int:
        async with get_db_context() as db:
            query = select(InvestigationTicket)

            conditions = []

            if employee_id:
                conditions.append(
                    or_(
                        InvestigationTicket.subject_employee_id == employee_id,
                    )
                )

            if department_id:
                conditions.append(InvestigationTicket.department_id == department_id)

            if assigned_officer_id:
                conditions.append(InvestigationTicket.assigned_officer_id == assigned_officer_id)

            if event_type:
                conditions.append(InvestigationTicket.event_type == event_type)

            if severity:
                conditions.append(InvestigationTicket.severity == severity)

            if status:
                conditions.append(InvestigationTicket.status == status)

            if date_from and date_to:
                field = getattr(InvestigationTicket.created_at)
                if date_field == "created_at" else getattr(InvestigationTicket.closed_at)
                conditions.append(field.between(date_from, date_to))
            elif date_from:
                field = getattr(InvestigationTicket, date_field, InvestigationTicket.created_at)
                conditions.append(field >= date_from)
            elif date_to:
                field = getattr(InvestigationTicket, date_field, InvestigationTicket.created_at)
                conditions.append(field <= date_to)

            if keyword:
                conditions.append(
                    or_(
                    InvestigationTicket.title.ilike(f"%{keyword}%"),
                    InvestigationTicket.description.ilike(f"%{keyword}%"),
                    InvestigationTicket.ticket_number.ilike(f"%{keyword}%"),
                    InvestigationTicket.subject_employee_name.ilike(f"%{keyword}%"),
                )
            )

            if conditions:
                query = query.where(and_(*conditions))

            count_query = select(func.count()).select_from(query.subquery())
            total_result = await db.execute(count_query)
            total = total_result.scalar() or 0

            if sort_order.lower() == "desc":
                query = query.order_by(desc(getattr(InvestigationTicket, sort_by, InvestigationTicket.created_at)))
            else:
                query = query.order_by(asc(getattr(InvestigationTicket, sort_by, InvestigationTicket.created_at)))

            offset = (page - 1) * page_size
            query = query.offset(offset).limit(page_size)

            result = await db.execute(query)
            tickets = result.scalars().all()

            return list(tickets), total

    async def query_compliance_events(
        self,
        employee_id: Optional[uuid.UUID] = None,
        event_type: Optional[str] = None,
        severity: Optional[str] = None,
        department_id: Optional[uuid.UUID] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        has_ticket: Optional[bool] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> Tuple[List[ComplianceEvent], int:
        async with get_db_context() as db:
            query = select(ComplianceEvent)
            conditions = []

            if employee_id:
                conditions.append(ComplianceEvent.subject_employee_id == employee_id)

            if department_id:
                conditions.append(ComplianceEvent.subject_department_id == department_id)

            if event_type:
                conditions.append(ComplianceEvent.event_type == event_type)

            if severity:
                conditions.append(ComplianceEvent.severity == severity)

            if date_from:
                conditions.append(ComplianceEvent.detected_at >= date_from)

            if date_to:
                conditions.append(ComplianceEvent.detected_at <= date_to)

            if has_ticket is not None:
                if has_ticket:
                    conditions.append(ComplianceEvent.ticket_id.isnot(None))
                else:
                    conditions.append(ComplianceEvent.ticket_id.is_(None))

            if conditions:
                query = query.where(and_(*conditions))

            count_query = select(func.count()).select_from(query.subquery())
            total_result = await db.execute(count_query)
            total = total_result.scalar() or 0

            query = query.order_by(desc(ComplianceEvent.detected_at))

            offset = (page - 1) * page_size
            query = query.offset(offset).limit(page_size)

            result = await db.execute(query)
            events = result.scalars().all()

            return list(events), total

    async def query_operation_logs(
        self,
        user_id: Optional[uuid.UUID] = None,
        action_type: Optional[str] = None,
        target_type: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        log_level: Optional[str] = None,
        page: int = 1,
        page_size: int = 100,
    ) -> Tuple[List[SystemLog], int:
        async with get_db_context() as db:
            query = select(SystemLog)
            conditions = []

            if user_id:
                conditions.append(SystemLog.user_id == user_id)

            if action_type:
                conditions.append(SystemLog.action_type.ilike(f"%{action_type}%"))

            if target_type:
                conditions.append(SystemLog.target_type == target_type)

            if date_from:
                conditions.append(SystemLog.created_at >= date_from)

            if date_to:
                conditions.append(SystemLog.created_at <= date_to)

            if log_level:
                conditions.append(SystemLog.log_level == log_level)

            if conditions:
                query = query.where(and_(*conditions))

            count_query = select(func.count()).select_from(query.subquery())
            total_result = await db.execute(count_query)
            total = total_result.scalar() or 0

            query = query.order_by(desc(SystemLog.created_at))

            offset = (page - 1) * page_size
            query = query.offset(offset).limit(page_size)

            result = await db.execute(query)
            logs = result.scalars().all()

            return list(logs), total

    async def query_compliance_profiles(
        self,
        employee_name: Optional[str] = None,
        department_id: Optional[uuid.UUID] = None,
        risk_level: Optional[str] = None,
        min_score: Optional[int] = None,
        max_score: Optional[int] = None,
        has_violations: Optional[bool] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> Tuple[List[ComplianceProfile], int:
        async with get_db_context() as db:
            query = select(ComplianceProfile).join(
                Employee, ComplianceProfile.employee_id == Employee.id
            )
            conditions = []

            if employee_name:
                conditions.append(Employee.name.ilike(f"%{employee_name}%"))

            if department_id:
                conditions.append(ComplianceProfile.department_id == department_id)

            if risk_level:
                conditions.append(ComplianceProfile.risk_level == risk_level)

            if min_score is not None:
                conditions.append(ComplianceProfile.compliance_score >= min_score)

            if max_score is not None:
                conditions.append(ComplianceProfile.compliance_score <= max_score)

            if has_violations is not None:
                if has_violations:
                    conditions.append(ComplianceProfile.confirmed_violations_count > 0)
                else:
                    conditions.append(ComplianceProfile.confirmed_violations_count == 0)

            if conditions:
                query = query.where(and_(*conditions))

            count_query = select(func.count()).select_from(query.subquery())
            total_result = await db.execute(count_query)
            total = total_result.scalar() or 0

            query = query.order_by(asc(ComplianceProfile.compliance_score))

            offset = (page - 1) * page_size
            query = query.offset(offset).limit(page_size)

            result = await db.execute(query)
            profiles = result.scalars().all()

            return list(profiles), total

    async def get_ticket_full_lifecycle(
        self, ticket_id: uuid.UUID
    ) -> Optional[Dict[str, Any]]:
        async with get_db_context() as db:
            ticket_result = await db.execute(
                select(InvestigationTicket).where(
                    InvestigationTicket.id == ticket_id
                )
            )
            ticket = ticket_result.scalar_one_or_none()
            if not ticket:
                return None

            events_result = await db.execute(
                select(ComplianceEvent).where(
                    ComplianceEvent.ticket_id == ticket_id
                )
            )
            events = events_result.scalars().all()

            evidence_result = await db.execute(
                select(EvidenceItem).where(
                    EvidenceItem.ticket_id == ticket_id
                )
            )
            evidence = evidence_result.scalars().all()

            from app.models.compliance import ApprovalRecord, EventTimeline

            approvals_result = await db.execute(
                select(ApprovalRecord).where(
                    ApprovalRecord.ticket_id == ticket_id
                )
            )
            approvals = approvals_result.scalars().all()

            timeline_result = await db.execute(
                select(EventTimeline).where(
                    EventTimeline.ticket_id == ticket_id
                ).order_by(EventTimeline.timestamp)
            )
            timeline = timeline_result.scalars().all()

            disciplinary_result = await db.execute(
                select(DisciplinaryActionRecord).where(
                    DisciplinaryActionRecord.ticket_id == ticket_id
                )
            )
            disciplinary = disciplinary_result.scalar_one_or_none()

            emp_result = await db.execute(
                select(Employee).where(
                    Employee.id == ticket.subject_employee_id
                )
            )
            employee = emp_result.scalar_one_or_none()

            return {
                "ticket": {
                    "id": str(ticket.id),
                    "ticket_number": ticket.ticket_number,
                    "title": ticket.title,
                    "description": ticket.description,
                    "event_type": ticket.event_type,
                    "severity": ticket.severity,
                    "status": ticket.status,
                    "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
                    "deadline": ticket.deadline.isoformat() if ticket.deadline else None,
                    "closed_at": ticket.closed_at.isoformat() if ticket.closed_at else None,
                    "priority_score": ticket.priority_score,
                    "subject_employee_id": str(ticket.subject_employee_id) if ticket.subject_employee_id else None,
                    "subject_employee_name": ticket.subject_employee_name,
                    "department_id": str(ticket.department_id) if ticket.department_id else None,
                    "department_name": ticket.department_name,
                    "assigned_officer_id": str(ticket.assigned_officer_id) if ticket.assigned_officer_id else None,
                    "assigned_officer_name": ticket.assigned_officer_name,
                    "violation_result": ticket.violation_result,
                    "conclusion_summary": ticket.conclusion_summary,
                    "disciplinary_action": ticket.disciplinary_action,
                    "is_overdue": ticket.is_overdue,
                    "escalation_count": ticket.escalation_count,
                },
                "employee_basic": {
                    "employee_id": str(employee.id) if employee else None,
                    "name": employee.name if employee else None,
                    "position": employee.position if employee else None,
                    "job_level": employee.job_level if employee else None,
                    "email": employee.email if employee else None,
                } if employee else None,
                "events": [
                    {
                        "id": str(e.id),
                        "event_code": e.event_code,
                        "event_type": e.event_type,
                        "severity": e.severity,
                        "detected_at": e.detected_at.isoformat() if e.detected_at else None,
                        "title": e.title,
                        "risk_score": e.risk_score,
                        "confidence": e.confidence,
                        "matched_data_sources": e.matched_data_sources,
                        "description": e.description,
                    }
                    for e in events
                ],
                "evidence_items": [
                    {
                        "id": str(ev.id),
                        "type": ev.evidence_type,
                        "data_source": ev.data_source,
                        "title": ev.title,
                        "description": ev.description,
                        "event_timestamp": ev.event_timestamp.isoformat() if ev.event_timestamp else None,
                        "is_key_evidence": ev.is_key_evidence,
                        "relevance_score": ev.relevance_score,
                    }
                    for ev in evidence
                ],
                "approvals": [
                    {
                        "id": str(a.id),
                        "approval_type": a.approval_type,
                        "order": a.approval_order,
                        "approver_name": a.approver_name,
                        "status": a.status,
                        "decision": a.decision,
                        "comments": a.comments,
                        "decided_at": a.decided_at.isoformat() if a.decided_at else None,
                        "deadline": a.deadline.isoformat() if a.deadline else None,
                    }
                    for a in approvals
                ],
                "timeline": [
                    {
                        "id": str(t.id),
                        "timestamp": t.timestamp.isoformat() if t.timestamp else None,
                        "type": t.timeline_type,
                        "title": t.title,
                        "description": t.description,
                        "is_key_event": t.is_key_event,
                        "employee_name": t.employee_name,
                    }
                    for t in timeline
                ],
                "disciplinary_action": {
                    "id": str(disciplinary.id) if disciplinary else None,
                    "action_type": disciplinary.action_type if disciplinary else None,
                    "status": disciplinary.status if disciplinary else None,
                    "effective_date": disciplinary.effective_date.isoformat() if disciplinary and disciplinary.effective_date else None,
                    "expiry_date": disciplinary.expiry_date.isoformat() if disciplinary and disciplinary.expiry_date else None,
                } if disciplinary else None,
            }

    async def export_tickets(
        self,
        query_params: Dict[str, Any],
        export_format: str = "xlsx",
        requester_id: Optional[uuid.UUID] = None,
    ) -> str:
        tickets, _ = await self.query_tickets(
            **query_params,
            page=1,
            page_size=self.MAX_EXPORT_ROWS,
        )

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"tickets_export_{timestamp}.{export_format}"
        filepath = os.path.join(settings.REPORT_OUTPUT_DIR, filename)

        os.makedirs(settings.REPORT_OUTPUT_DIR, exist_ok=True)

        cn_status = {
            "pending": "待分配", "assigned": "已分配",
            "under_investigation": "调查中", "evidence_collected": "已取证",
            "conclusion_submitted": "待审批",
            "under_approval": "审批中",
            "approved": "审批通过", "rejected": "审批驳回",
            "closed": "已关闭", "escalated": "已升级",
        }
        cn_severity = {"critical": "重大", "important": "重要", "general": "一般"}

        headers = [
            "工单号", "标题", "事件类型", "严重程度", "当前状态",
            "涉事员工", "所属部门", "指派专员",
            "创建时间", "截止时间", "关闭时间",
            "违规认定", "纪律处分", "是否逾期",
            "风险评分", "完成时长(小时)"
        ]

        if export_format == "csv":
            with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for t in tickets:
                    writer.writerow([
                        t.ticket_number,
                        t.title,
                        t.event_type,
                        cn_severity.get(t.severity, t.severity),
                        cn_status.get(t.status, t.status),
                        t.subject_employee_name or "",
                        t.department_name or "",
                        t.assigned_officer_name or "",
                        t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                        t.deadline.strftime("%Y-%m-%d %H:%M:%S") if t.deadline else "",
                        t.closed_at.strftime("%Y-%m-%d %H:%M:%S") if t.closed_at else "",
                        t.violation_result or "",
                        t.disciplinary_action or "",
                        "是" if t.is_overdue else "否",
                        t.priority_score,
                        round((t.actual_hours or 0), 2),
                    ])
        else:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工单数据"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            for row_idx, t in enumerate(tickets, 2):
                values = [
                    t.ticket_number,
                    t.title,
                    t.event_type,
                    cn_severity.get(t.severity, t.severity),
                    cn_status.get(t.status, t.status),
                    t.subject_employee_name or "",
                    t.department_name or "",
                    t.assigned_officer_name or "",
                    t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                    t.deadline.strftime("%Y-%m-%d %H:%M:%S") if t.deadline else "",
                    t.closed_at.strftime("%Y-%m-%d %H:%M:%S") if t.closed_at else "",
                    t.violation_result or "",
                    t.disciplinary_action or "",
                    "是" if t.is_overdue else "否",
                    t.priority_score,
                    round((t.actual_hours or 0), 2),
                ]
                for col, val in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col, value=val)

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = [
                    22, 50, 18, 12, 12, 18, 20, 18, 20, 20, 20, 15, 15, 10, 12, 15
                ][col - 1]

            wb.save(filepath)

        async with get_db_context() as db:
            log = SystemLog(
                id=uuid.uuid4(),
                log_level="INFO",
                action_type=LogActionType.DATA_EXPORT.value,
                target_type="export_file",
                target_name=filename,
                user_id=requester_id,
                action_details={
                    "format": export_format,
                    "record_count": len(tickets),
                    "query_params": {k: str(v) for k, v in query_params.items()},
                    "file_path": filepath,
                },
                status="success",
            )
            db.add(log)

        self.logger.info(
            "Data export completed",
            format=export_format,
            filename=filename,
            records=len(tickets),
            user=str(requester_id)
        )

        return filepath

    async def export_operation_logs(
        self,
        query_params: Dict[str, Any],
        requester_id: Optional[uuid.UUID] = None,
    ) -> str:
        logs, _ = await self.query_operation_logs(
            **query_params,
            page=1,
            page_size=self.MAX_EXPORT_ROWS,
        )

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"operation_logs_{timestamp}.csv"
        filepath = os.path.join(settings.REPORT_OUTPUT_DIR, filename)
        os.makedirs(settings.REPORT_OUTPUT_DIR, exist_ok=True)

        headers = [
            "时间", "操作类型", "操作人", "目标类型", "目标名称",
            "日志级别", "状态", "IP地址", "操作详情"
        ]

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for log in logs:
                writer.writerow([
                    log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
                    log.action_type,
                    log.user_name or "",
                    log.target_type or "",
                    log.target_name or "",
                    log.log_level,
                    log.status,
                    log.ip_address or "",
                    json.dumps(log.action_details, ensure_ascii=False)[:500] if log.action_details else "",
                ])

        async with get_db_context() as db:
            sys_log = SystemLog(
                id=uuid.uuid4(),
                log_level="INFO",
                action_type=LogActionType.DATA_EXPORT.value,
                target_type="export_file",
                target_name=filename,
                user_id=requester_id,
                action_details={
                    "format": "csv",
                    "record_count": len(logs),
                    "file_path": filepath,
                },
                status="success",
            )
            db.add(sys_log)

        self.logger.info(
            "Operation logs exported",
            filename=filename,
            records=len(logs)
        )

        return filepath

    async def get_statistics_dashboard(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        if not date_from:
            date_from = datetime.utcnow() - timedelta(days=30)
        if not date_to:
            date_to = datetime.utcnow()

        async with get_db_context() as db:
            stats = {}

            ticket_overview = await db.execute(
                select(
                    func.count(InvestigationTicket.id),
                    func.sum(
                        case(
                        (InvestigationTicket.status == TicketStatus.CLOSED.value, 1),
                        else_=0
                    )
                ),
                ).where(
                    InvestigationTicket.created_at.between(date_from, date_to)
                )
            )
            total_created, total_closed = ticket_overview.one()

            stats["ticket_overview"] = {
                "total_created": total_created or 0,
                "total_closed": total_closed or 0,
                "pending": (total_created or 0) - (total_closed or 0),
                "completion_rate": round(
                    (total_closed or 0) / (total_created or 1) * 100, 2
                ),
            }

            sev_result = await db.execute(
                select(
                    InvestigationTicket.severity,
                    func.count(InvestigationTicket.id)
                ).where(
                    InvestigationTicket.created_at.between(date_from, date_to)
                ).group_by(InvestigationTicket.severity)
            )
            stats["tickets_by_severity"] = dict(sev_result.all())

            type_result = await db.execute(
                select(
                    InvestigationTicket.event_type,
                    func.count(InvestigationTicket.id)
                ).where(
                    InvestigationTicket.created_at.between(date_from, date_to)
                ).group_by(InvestigationTicket.event_type)
            )
            stats["tickets_by_type"] = dict(type_result.all())

            dept_result = await db.execute(
                select(
                    InvestigationTicket.department_name,
                    func.count(InvestigationTicket.id)
                ).where(
                    and_(
                        InvestigationTicket.created_at.between(date_from, date_to),
                        InvestigationTicket.department_name.isnot(None)
                    )
                ).group_by(InvestigationTicket.department_name)
                .order_by(desc(func.count(InvestigationTicket.id)))
                .limit(10)
            )
            stats["top_departments"] = [
                {"name": name or "未知", "count": count}
                for name, count in dept_result.all()
            ]

            result_distribution = await db.execute(
                select(
                    InvestigationTicket.violation_result,
                    func.count(InvestigationTicket.id)
                ).where(
                    and_(
                        InvestigationTicket.status == TicketStatus.CLOSED.value,
                        InvestigationTicket.closed_at.between(date_from, date_to),
                        InvestigationTicket.violation_result.isnot(None)
                    )
                ).group_by(InvestigationTicket.violation_result)
            )
            stats["result_distribution"] = dict(result_distribution.all())

            overdue_count = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    InvestigationTicket.is_overdue == True
                )
            )
            stats["overdue_count"] = overdue_count.scalar() or 0

            pending_approvals = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    InvestigationTicket.status == TicketStatus.UNDER_APPROVAL.value
                )
            )
            stats["pending_approvals"] = pending_approvals.scalar() or 0

            avg_processing = await db.execute(
                select(func.avg(
                    func.extract("epoch", InvestigationTicket.closed_at - InvestigationTicket.created_at) / 3600
                )).where(
                    and_(
                        InvestigationTicket.status == TicketStatus.CLOSED.value,
                        InvestigationTicket.closed_at.between(date_from, date_to)
                    )
                )
            )
            stats["avg_processing_hours"] = round(
                avg_processing.scalar() or 0, 2
            )

            active_officers = await db.execute(
                select(
                    Employee.name,
                    func.count(InvestigationTicket.id)
                ).join(
                    InvestigationTicket.assigned_officer_id == Employee.id
                ).where(
                    InvestigationTicket.status.in_([
                        TicketStatus.ASSIGNED.value,
                        TicketStatus.UNDER_INVESTIGATION.value,
                    ])
                ).group_by(Employee.name
                ).order_by(desc(func.count(InvestigationTicket.id)))
            )
            stats["officer_workload"] = [
                {"name": name, "active_count": count}
                for name, count in active_officers.all()
            ]

            return stats
