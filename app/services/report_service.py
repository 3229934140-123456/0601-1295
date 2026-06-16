from __future__ import annotations
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
import uuid
import os
import json
from collections import defaultdict, Counter
import io
from sqlalchemy import select, and_, or_, func, extract
from sqlalchemy.ext.asyncio import AsyncSession
from app.core import logger, settings
from app.core.constants import (
    SeverityLevel, TicketStatus, ViolationResult,
    EventType, DisciplinaryAction, LogActionType
)
from app.core.database import get_db_context
from app.models.investigation import InvestigationTicket, ComplianceEvent
from app.models.compliance import DailyStatistics, SystemLog, EmployeeReport
from app.models.organization import Employee, InvestigationOfficer, Department


class ReportService:
    def __init__(self):
        self.logger = logger.bind(module="ReportService")
        os.makedirs(settings.REPORT_OUTPUT_DIR, exist_ok=True)

    async def generate_daily_report(
        self, report_date: Optional[datetime] = None
    ) -> Tuple[str, DailyStatistics]:
        report_date = report_date or datetime.utcnow()
        start_of_day = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)

        async with get_db_context() as db:
            stats = await self._compute_daily_statistics(db, start_of_day, end_of_day)
            await db.flush()

            pdf_path = await self._generate_pdf_report(db, stats, start_of_day)
            excel_path = await self._generate_excel_report(db, stats, start_of_day)

            stats.reports_generated = (stats.reports_generated or 0) + 2

            log = SystemLog(
                id=uuid.uuid4(),
                log_level="INFO",
                action_type=LogActionType.REPORT_GENERATED.value,
                target_type="daily_report",
                target_id=stats.id,
                target_name=f"DAILY-{start_of_day.strftime('%Y%m%d')}",
                action_details={
                    "pdf_path": pdf_path,
                    "excel_path": excel_path,
                    "report_date": start_of_day.isoformat(),
                },
                status="success",
            )
            db.add(log)

            self.logger.info(
                "Daily report generated",
                date=start_of_day.strftime("%Y-%m-%d"),
                pdf_path=pdf_path,
                excel_path=excel_path
            )
            return (pdf_path, excel_path), stats

    async def _compute_daily_statistics(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> DailyStatistics:
        existing = await db.execute(
            select(DailyStatistics).where(DailyStatistics.stat_date == start)
        )
        stats = existing.scalar_one_or_none()

        if not stats:
            stats = DailyStatistics(
                id=uuid.uuid4(),
                stat_date=start,
            )
            db.add(stats)
            await db.flush()

        stats.total_data_collected = await self._count_data_collected(db, start, end)
        stats.data_by_source = await self._count_data_by_source(db, start, end)

        events_result = await db.execute(
            select(ComplianceEvent).where(
                ComplianceEvent.detected_at.between(start, end)
            )
        )
        events = events_result.scalars().all()
        stats.total_events_detected = len(events)
        stats.events_by_type = dict(Counter(e.event_type for e in events))
        stats.events_by_severity = dict(Counter(e.severity for e in events))
        stats.duplicate_events = sum(1 for e in events if e.is_duplicate)

        tickets_created_result = await db.execute(
            select(InvestigationTicket).where(
                InvestigationTicket.created_at.between(start, end)
            )
        )
        tickets_created = tickets_created_result.scalars().all()
        stats.total_tickets_created = len(tickets_created)
        stats.tickets_created_by_severity = dict(Counter(t.severity for t in tickets_created))

        closed_tickets_result = await db.execute(
            select(InvestigationTicket).where(
                and_(
                    InvestigationTicket.closed_at.isnot(None),
                    InvestigationTicket.closed_at.between(start, end)
                )
            )
        )
        closed_tickets = closed_tickets_result.scalars().all()
        stats.total_tickets_closed = len(closed_tickets)
        stats.tickets_closed_by_result = dict(
            Counter(t.violation_result for t in closed_tickets if t.violation_result)
        )
        stats.tickets_closed_by_type = dict(Counter(t.event_type for t in closed_tickets))

        pending_result = await db.execute(
            select(InvestigationTicket).where(
                InvestigationTicket.status.in_([
                    TicketStatus.PENDING.value,
                    TicketStatus.ASSIGNED.value,
                    TicketStatus.UNDER_INVESTIGATION.value,
                    TicketStatus.EVIDENCE_COLLECTED.value,
                    TicketStatus.CONCLUSION_SUBMITTED.value,
                    TicketStatus.UNDER_APPROVAL.value,
                    TicketStatus.ESCALATED.value,
                ])
            )
        )
        pending = pending_result.scalars().all()
        stats.pending_tickets = len(pending)
        stats.overdue_tickets = sum(1 for t in pending if t.is_overdue)
        stats.escalated_tickets = sum(
            1 for t in pending if t.status == TicketStatus.ESCALATED.value
        )

        approval_result = await db.execute(
            select(InvestigationTicket).where(
                InvestigationTicket.status == TicketStatus.UNDER_APPROVAL.value
            )
        )
        stats.tickets_in_approval = len(approval_result.scalars().all())

        reports_result = await db.execute(
            select(EmployeeReport).where(
                EmployeeReport.created_at.between(start, end)
            )
        )
        reports = reports_result.scalars().all()
        stats.employee_reports_submitted = len(reports)
        stats.employee_reports_merged = sum(
            1 for r in reports if r.status == "merged"
        )

        stats.confirmed_violations = sum(
            1 for t in closed_tickets
            if t.violation_result == ViolationResult.CONFIRMED.value
        )
        stats.false_positives = sum(
            1 for t in closed_tickets
            if t.violation_result == ViolationResult.FALSE_POSITIVE.value
        )

        disciplinary_actions = []
        for t in closed_tickets:
            if t.disciplinary_action and t.violation_result == ViolationResult.CONFIRMED.value:
                disciplinary_actions.append(t.disciplinary_action)
        stats.disciplinary_actions_issued = dict(Counter(disciplinary_actions))

        avg_hours, avg_by_severity, avg_by_type = self._calculate_avg_processing(
            closed_tickets
        )
        stats.avg_processing_hours = avg_hours
        stats.avg_processing_hours_by_severity = avg_by_severity
        stats.avg_processing_hours_by_type = avg_by_type

        total_tickets = stats.total_tickets_closed + stats.pending_tickets
        stats.completion_rate = (
            stats.total_tickets_closed / total_tickets if total_tickets > 0 else 1.0
        )
        on_time_closed = sum(
            1 for t in closed_tickets if not t.is_overdue
        )
        stats.on_time_rate = (
            on_time_closed / stats.total_tickets_closed
            if stats.total_tickets_closed > 0 else 1.0
        )

        stats.officer_workload = await self._compute_officer_workload(db)
        stats.department_statistics = await self._compute_department_stats(db, start, end)

        top_events = Counter(e.event_type for e in events).most_common(5)
        stats.top_event_types = [
            {"type": t, "count": c, "label": self._event_type_label(t)}
            for t, c in top_events
        ]

        stats.trend_7days = await self._compute_trend(db, start, 7)
        stats.trend_30days = await self._compute_trend(db, start, 30)

        notifications_log = await db.execute(
            select(SystemLog).where(
                and_(
                    SystemLog.created_at.between(start, end),
                    SystemLog.action_type == "notification_sent"
                )
            )
        )
        stats.notifications_sent = len(notifications_log.scalars().all())

        return stats

    @staticmethod
    async def _count_data_collected(
        db: AsyncSession, start: datetime, end: datetime
    ) -> int:
        from app.models.data_source import RawDataRecord
        result = await db.execute(
            select(func.count(RawDataRecord.id)).where(
                RawDataRecord.collected_at.between(start, end)
            )
        )
        return result.scalar() or 0

    @staticmethod
    async def _count_data_by_source(
        db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, int]:
        from app.models.data_source import RawDataRecord
        result = await db.execute(
            select(
                RawDataRecord.data_source,
                func.count(RawDataRecord.id)
            ).where(
                RawDataRecord.collected_at.between(start, end)
            ).group_by(RawDataRecord.data_source)
        )
        return dict(result.all())

    @staticmethod
    def _calculate_avg_processing(
        closed_tickets: List[InvestigationTicket]
    ) -> Tuple[Dict, Dict, Dict]:
        overall = []
        by_severity: Dict[str, List] = defaultdict(list)
        by_type: Dict[str, List] = defaultdict(list)

        for t in closed_tickets:
            if t.closed_at and t.created_at:
                hours = (t.closed_at - t.created_at).total_seconds() / 3600
                overall.append(hours)
                by_severity[t.severity].append(hours)
                by_type[t.event_type].append(hours)

        avg = lambda lst: round(sum(lst) / len(lst), 2) if lst else 0
        return (
            {"overall": avg(overall), "count": len(overall)},
            {k: {"avg": avg(v), "count": len(v)} for k, v in by_severity.items()},
            {k: {"avg": avg(v), "count": len(v)} for k, v in by_type.items()},
        )

    @staticmethod
    async def _compute_officer_workload(
        db: AsyncSession
    ) -> List[Dict[str, Any]]:
        officers_result = await db.execute(select(InvestigationOfficer))
        officers = officers_result.scalars().all()
        workload = []

        for officer in officers:
            if not officer.employee_id:
                continue

            active_result = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    and_(
                        InvestigationTicket.assigned_officer_id == officer.employee_id,
                        InvestigationTicket.status.in_([
                            TicketStatus.ASSIGNED.value,
                            TicketStatus.UNDER_INVESTIGATION.value,
                            TicketStatus.EVIDENCE_COLLECTED.value,
                        ])
                    )
                )
            )
            active = active_result.scalar() or 0

            closed_result = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    and_(
                        InvestigationTicket.assigned_officer_id == officer.employee_id,
                        InvestigationTicket.status == TicketStatus.CLOSED.value,
                        InvestigationTicket.closed_at >= datetime.utcnow() - timedelta(days=30)
                    )
                )
            )
            closed_30d = closed_result.scalar() or 0

            emp_result = await db.execute(
                select(Employee).where(Employee.id == officer.employee_id)
            )
            emp = emp_result.scalar_one_or_none()

            workload.append({
                "officer_id": str(officer.employee_id),
                "name": emp.name if emp else "未知",
                "specializations": officer.specializations or [],
                "active_tickets": active,
                "capacity": officer.max_ticket_capacity,
                "closed_last_30d": closed_30d,
                "utilization_rate": round(
                    active / officer.max_ticket_capacity * 100, 1
                ) if officer.max_ticket_capacity > 0 else 0,
            })

        return sorted(workload, key=lambda x: -x["utilization_rate"])

    @staticmethod
    async def _compute_department_stats(
        db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        depts_result = await db.execute(select(Department))
        departments = depts_result.scalars().all()
        dept_stats = []

        for dept in departments:
            tickets_result = await db.execute(
                select(InvestigationTicket).where(
                    InvestigationTicket.department_id == dept.id
                )
            )
            dept_tickets = tickets_result.scalars().all()

            total = len(dept_tickets)
            active = sum(1 for t in dept_tickets if t.status != TicketStatus.CLOSED.value)
            closed = total - active
            confirmed = sum(
                1 for t in dept_tickets
                if t.violation_result == ViolationResult.CONFIRMED.value
            )

            employees_result = await db.execute(
                select(func.count(Employee.id)).where(
                    Employee.department_id == dept.id,
                    Employee.employment_status == "active"
                )
            )
            emp_count = employees_result.scalar() or 0

            rate = round((confirmed / emp_count * 1000), 2) if emp_count > 0 else 0

            dept_stats.append({
                "department_id": str(dept.id),
                "name": dept.name,
                "employee_count": emp_count,
                "total_tickets": total,
                "active_tickets": active,
                "closed_tickets": closed,
                "confirmed_violations": confirmed,
                "violation_rate_per_1000": rate,
                "severity_breakdown": dict(Counter(t.severity for t in dept_tickets)),
            })

        return sorted(dept_stats, key=lambda x: -x["confirmed_violations"])

    async def _compute_trend(
        self, db: AsyncSession, end_date: datetime, days: int
    ) -> Dict[str, Any]:
        trend_data = {
            "dates": [],
            "events_count": [],
            "tickets_created": [],
            "tickets_closed": [],
            "by_severity": {
                SeverityLevel.CRITICAL.value: [],
                SeverityLevel.IMPORTANT.value: [],
                SeverityLevel.GENERAL.value: [],
            }
        }

        for i in range(days - 1, -1, -1):
            day_start = (end_date - timedelta(days=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            day_end = day_start + timedelta(days=1)
            date_label = day_start.strftime("%m-%d")
            trend_data["dates"].append(date_label)

            events_count = await db.execute(
                select(func.count(ComplianceEvent.id)).where(
                    ComplianceEvent.detected_at.between(day_start, day_end)
                )
            )
            trend_data["events_count"].append(events_count.scalar() or 0)

            tickets_created = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    InvestigationTicket.created_at.between(day_start, day_end)
                )
            )
            trend_data["tickets_created"].append(tickets_created.scalar() or 0)

            tickets_closed = await db.execute(
                select(func.count(InvestigationTicket.id)).where(
                    and_(
                        InvestigationTicket.closed_at.isnot(None),
                        InvestigationTicket.closed_at.between(day_start, day_end)
                    )
                )
            )
            trend_data["tickets_closed"].append(tickets_closed.scalar() or 0)

            for severity in [SeverityLevel.CRITICAL, SeverityLevel.IMPORTANT, SeverityLevel.GENERAL]:
                sev_count = await db.execute(
                    select(func.count(ComplianceEvent.id)).where(
                        and_(
                            ComplianceEvent.detected_at.between(day_start, day_end),
                            ComplianceEvent.severity == severity.value
                        )
                    )
                )
                trend_data["by_severity"][severity.value].append(sev_count.scalar() or 0)

        return trend_data

    async def _generate_pdf_report(
        self, db: AsyncSession, stats: DailyStatistics, report_date: datetime
    ) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            )
            from reportlab.graphics.shapes import Drawing, Line
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics import renderPDF

            filename = f"合规日报_{report_date.strftime('%Y%m%d')}.pdf"
            filepath = os.path.join(settings.REPORT_OUTPUT_DIR, filename)

            doc = SimpleDocTemplate(
                filepath,
                pagesize=A4,
                topMargin=20*mm,
                bottomMargin=20*mm,
                leftMargin=15*mm,
                rightMargin=15*mm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a365d'),
                alignment=1,
                spaceAfter=10
            )
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#4a5568'),
                alignment=1,
                spaceAfter=20
            )
            h2_style = ParagraphStyle(
                'H2',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#2d3748'),
                spaceBefore=15,
                spaceAfter=8
            )

            story = []

            story.append(Paragraph("企业合规管理系统 - 每日报告", title_style))
            story.append(Paragraph(
                f"报告日期: {report_date.strftime('%Y年%m月%d日')} &nbsp;&nbsp;&nbsp; "
                f"生成时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
                subtitle_style
            ))

            overview_data = [
                ["指标名称", "数值", "说明"],
                ["数据采集总量", f"{stats.total_data_collected:,}", "各数据源采集记录数"],
                ["检测到违规事件", f"{stats.total_events_detected:,}", "规则引擎检测结果"],
                ["重复事件过滤", f"{stats.duplicate_events}", "去重机制过滤数量"],
                ["新生成工单", f"{stats.total_tickets_created:,}", "调查工单数量"],
                ["结案件数", f"{stats.total_tickets_closed:,}", "完成处理工单"],
                ["待处理工单", f"{stats.pending_tickets:,}", "仍在流程中的工单"],
                ["逾期工单", f"{stats.overdue_tickets:,}", "超时限未处理"],
                ["升级工单", f"{stats.escalated_tickets:,}", "自动升级至主管"],
                ["员工主动申报", f"{stats.employee_reports_submitted}", "员工举报/申报数"],
                ["完成率", f"{stats.completion_rate * 100:.1f}%", "工单按期完成率"],
                ["准时率", f"{stats.on_time_rate * 100:.1f}%", "按时完成比例"],
            ]
            overview_table = Table(overview_data, colWidths=[55*mm, 35*mm, 75*mm])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(Paragraph("一、总体概览", h2_style))
            story.append(overview_table)

            story.append(Paragraph("二、事件类型分布 TOP 5", h2_style))
            if stats.top_event_types:
                event_data = [["排名", "事件类型", "数量", "占比"]]
                total_evt = sum(t["count"] for t in stats.top_event_types)
                for idx, t in enumerate(stats.top_event_types, 1):
                    pct = f"{t['count'] / total_evt * 100:.1f}%" if total_evt > 0 else "0%"
                    event_data.append([
                        str(idx), t["label"],
                        str(t["count"]), pct
                    ])
                evt_table = Table(event_data, colWidths=[20*mm, 60*mm, 30*mm, 30*mm])
                evt_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#48bb78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                    ('PADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(evt_table)

            story.append(Paragraph("三、严重程度分布", h2_style))
            sev_data = [["严重程度", "事件数", "工单数"]]
            cn_map = {"critical": "重大", "important": "重要", "general": "一般"}
            for sev, cn in cn_map.items():
                evts = stats.events_by_severity.get(sev, 0)
                tkts = stats.tickets_created_by_severity.get(sev, 0)
                sev_data.append([cn, str(evts), str(tkts)])
            sev_table = Table(sev_data, colWidths=[40*mm, 50*mm, 50*mm])
            sev_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ed8936')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(sev_table)

            story.append(Paragraph("四、处理结果统计", h2_style))
            result_data = [["处理结果", "工单数", "占比"]]
            for result, cnt in stats.tickets_closed_by_result.items():
                pct = f"{cnt / stats.total_tickets_closed * 100:.1f}%" if stats.total_tickets_closed > 0 else "0%"
                result_cn = {
                    "confirmed": "确认违规",
                    "unconfirmed": "无法确认",
                    "false_positive": "误报排除",
                    "pending": "待认定"
                }.get(result, result)
                result_data.append([result_cn, str(cnt), pct])
            if len(result_data) == 1:
                result_data.append(["（暂无数据）", "-", "-"])
            result_table = Table(result_data, colWidths=[50*mm, 40*mm, 40*mm])
            result_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9f7aea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(result_table)

            if stats.disciplinary_actions_issued:
                story.append(Paragraph("五、纪律处分统计", h2_style))
                action_data = [["处分类型", "执行次数"]]
                action_cn = {
                    "warning": "警告",
                    "serious_warning": "记过",
                    "demotion": "降级",
                    "salary_reduction": "降薪",
                    "permission_freeze": "冻结权限",
                    "termination": "解除合同",
                    "training": "合规培训",
                    "no_action": "免于处罚"
                }
                for action, cnt in sorted(stats.disciplinary_actions_issued.items(), key=lambda x: -x[1]):
                    action_data.append([action_cn.get(action, action), str(cnt)])
                act_table = Table(action_data, colWidths=[70*mm, 50*mm])
                act_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e53e3e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                    ('PADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(act_table)

            if stats.avg_processing_hours:
                story.append(Paragraph("六、处理时效分析", h2_style))
                avg_data = [["指标", "平均处理时长(小时)", "样本数"]]
                overall = stats.avg_processing_hours or {}
                avg_data.append([
                    "全部工单",
                    f"{overall.get('overall', 0):.2f}",
                    str(overall.get('count', 0))
                ])
                for sev, info in (stats.avg_processing_hours_by_severity or {}).items():
                    sev_cn = {"critical": "重大", "important": "重要", "general": "一般"}.get(sev, sev)
                    avg_data.append([
                        f"{sev_cn}级",
                        f"{info.get('avg', 0):.2f}",
                        str(info.get('count', 0))
                    ])
                avg_table = Table(avg_data, colWidths=[45*mm, 55*mm, 35*mm])
                avg_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38b2ac')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                    ('PADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(avg_table)

            story.append(Spacer(1, 10*mm))
            story.append(Paragraph(
                f"<i>本报告由合规管理系统自动生成，数据截止至 {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}。</i>",
                styles['Italic']
            ))

            doc.build(story)
            return filepath

        except Exception as e:
            self.logger.error("Failed to generate PDF report", error=str(e))
            fallback_path = os.path.join(
                settings.REPORT_OUTPUT_DIR,
                f"合规日报_{report_date.strftime('%Y%m%d')}_data.json"
            )
            with open(fallback_path, "w", encoding="utf-8") as f:
                json.dump({
                    "report_date": report_date.isoformat(),
                    "stats": {
                        "total_data_collected": stats.total_data_collected,
                        "total_events_detected": stats.total_events_detected,
                        "total_tickets_created": stats.total_tickets_created,
                        "total_tickets_closed": stats.total_tickets_closed,
                        "pending_tickets": stats.pending_tickets,
                        "overdue_tickets": stats.overdue_tickets,
                    },
                    "error": str(e),
                }, f, ensure_ascii=False, indent=2, default=str)
            return fallback_path

    async def _generate_excel_report(
        self, db: AsyncSession, stats: DailyStatistics, report_date: datetime
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            filename = f"合规日报_{report_date.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(settings.REPORT_OUTPUT_DIR, filename)

            wb = openpyxl.Workbook()

            header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            alt_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            def style_header(ws, row, max_col):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

            def style_data(ws, start_row, end_row, max_col):
                for row in range(start_row, end_row + 1):
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = center_align
                        cell.border = thin_border
                        if (row - start_row) % 2 == 1:
                            cell.fill = alt_fill

            ws1 = wb.active
            ws1.title = "总体概览"
            ws1['A1'] = "企业合规管理系统 - 每日统计报告"
            ws1['A1'].font = Font(name='Microsoft YaHei', bold=True, size=16, color='1F4E79')
            ws1.merge_cells('A1:D1')
            ws1['A2'] = f"报告日期: {report_date.strftime('%Y-%m-%d')}"
            ws1['A2'].font = Font(name='Microsoft YaHei', size=10, color='666666')
            ws1.merge_cells('A2:D2')

            headers = ["指标类别", "指标名称", "数值", "备注"]
            for col, h in enumerate(headers, 1):
                ws1.cell(row=4, column=col, value=h)
            style_header(ws1, 4, 4)

            overview_rows = [
                ["数据采集", "采集总记录数", stats.total_data_collected, "各数据源合计"],
                ["数据采集", "邮件记录", stats.data_by_source.get("email", 0), "邮件系统同步"],
                ["数据采集", "即时通讯记录", stats.data_by_source.get("instant_message", 0), "聊天记录同步"],
                ["数据采集", "门禁记录", stats.data_by_source.get("door_access", 0), "门禁系统同步"],
                ["数据采集", "财务报销记录", stats.data_by_source.get("finance", 0), "财务系统同步"],
                ["违规检测", "检测到事件总数", stats.total_events_detected, "规则引擎检测"],
                ["违规检测", "重大事件", stats.events_by_severity.get("critical", 0), "红色预警"],
                ["违规检测", "重要事件", stats.events_by_severity.get("important", 0), "橙色预警"],
                ["违规检测", "一般事件", stats.events_by_severity.get("general", 0), "黄色提示"],
                ["违规检测", "去重过滤数", stats.duplicate_events, "智能去重机制"],
                ["工单处理", "新生成工单", stats.total_tickets_created, "自动生成+人工创建"],
                ["工单处理", "今日结案数", stats.total_tickets_closed, "完成全流程"],
                ["工单处理", "待处理工单", stats.pending_tickets, "进行中"],
                ["工单处理", "逾期工单", stats.overdue_tickets, "超时限"],
                ["工单处理", "审批中工单", stats.tickets_in_approval, "多级审批"],
                ["工单处理", "已升级工单", stats.escalated_tickets, "主管介入"],
                ["工单处理", "按时完成率", f"{stats.on_time_rate * 100:.1f}%", "时效指标"],
                ["工单处理", "总体完成率", f"{stats.completion_rate * 100:.1f}%", "进度指标"],
                ["处理结果", "确认违规", stats.confirmed_violations, "纪律处分"],
                ["处理结果", "误报排除", stats.false_positives, "规则优化参考"],
                ["员工参与", "主动申报", stats.employee_reports_submitted, "员工举报/申报"],
                ["员工参与", "已合并处理", stats.employee_reports_merged, "并入现有工单"],
            ]

            for row_idx, row_data in enumerate(overview_rows, 5):
                for col_idx, val in enumerate(row_data, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=val)
            style_data(ws1, 5, 5 + len(overview_rows) - 1, 4)

            for col in [1, 2, 3, 4]:
                ws1.column_dimensions[get_column_letter(col)].width = [18, 25, 18, 30][col-1]

            ws2 = wb.create_sheet("按部门统计")
            ws2['A1'] = "各部门合规情况统计"
            ws2['A1'].font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
            ws2.merge_cells('A1:H1')

            dept_headers = ["部门名称", "员工数", "总工单数", "进行中", "已结案",
                          "确认违规", "违规率(‰)", "严重程度分布"]
            for col, h in enumerate(dept_headers, 1):
                ws2.cell(row=3, column=col, value=h)
            style_header(ws2, 3, 8)

            for row_idx, dept in enumerate(stats.department_statistics or [], 4):
                sev_dist = dept.get("severity_breakdown", {})
                dist_str = f"重大:{sev_dist.get('critical',0)} 重要:{sev_dist.get('important',0)} 一般:{sev_dist.get('general',0)}"
                ws2.cell(row=row_idx, column=1, value=dept.get("name", ""))
                ws2.cell(row=row_idx, column=2, value=dept.get("employee_count", 0))
                ws2.cell(row=row_idx, column=3, value=dept.get("total_tickets", 0))
                ws2.cell(row=row_idx, column=4, value=dept.get("active_tickets", 0))
                ws2.cell(row=row_idx, column=5, value=dept.get("closed_tickets", 0))
                ws2.cell(row=row_idx, column=6, value=dept.get("confirmed_violations", 0))
                ws2.cell(row=row_idx, column=7, value=dept.get("violation_rate_per_1000", 0))
                ws2.cell(row=row_idx, column=8, value=dist_str)

            if stats.department_statistics:
                style_data(ws2, 4, 4 + len(stats.department_statistics) - 1, 8)
            for col in range(1, 9):
                ws2.column_dimensions[get_column_letter(col)].width = [25, 12, 12, 12, 12, 12, 15, 35][col-1]

            ws3 = wb.create_sheet("调查专员负荷")
            ws3['A1'] = "调查专员工作负荷统计"
            ws3['A1'].font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
            ws3.merge_cells('A1:F1')

            officer_headers = ["专员姓名", "专长领域", "在办工单数", "最大容量", "近30日结案", "利用率(%)"]
            for col, h in enumerate(officer_headers, 1):
                ws3.cell(row=3, column=col, value=h)
            style_header(ws3, 3, 6)

            for row_idx, officer in enumerate(stats.officer_workload or [], 4):
                specs = ", ".join(officer.get("specializations", []))
                ws3.cell(row=row_idx, column=1, value=officer.get("name", ""))
                ws3.cell(row=row_idx, column=2, value=specs)
                ws3.cell(row=row_idx, column=3, value=officer.get("active_tickets", 0))
                ws3.cell(row=row_idx, column=4, value=officer.get("capacity", 0))
                ws3.cell(row=row_idx, column=5, value=officer.get("closed_last_30d", 0))
                ws3.cell(row=row_idx, column=6, value=officer.get("utilization_rate", 0))

            if stats.officer_workload:
                style_data(ws3, 4, 4 + len(stats.officer_workload) - 1, 6)
            for col in range(1, 7):
                ws3.column_dimensions[get_column_letter(col)].width = [18, 30, 15, 15, 18, 15][col-1]

            ws4 = wb.create_sheet("趋势数据")
            trend = stats.trend_30days or {}
            if trend:
                dates = trend.get("dates", [])
                ws4['A1'] = "近30日趋势数据"
                ws4['A1'].font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
                ws4.merge_cells(f'A1:F1')

                trend_headers = ["日期", "违规事件数", "新增工单", "结案数", "重大事件", "重要事件"]
                for col, h in enumerate(trend_headers, 1):
                    ws4.cell(row=3, column=col, value=h)
                style_header(ws4, 3, 6)

                critical = trend.get("by_severity", {}).get("critical", [])
                important = trend.get("by_severity", {}).get("important", [])
                events = trend.get("events_count", [])
                created = trend.get("tickets_created", [])
                closed = trend.get("tickets_closed", [])

                for i in range(len(dates)):
                    row = 4 + i
                    ws4.cell(row=row, column=1, value=dates[i])
                    ws4.cell(row=row, column=2, value=events[i] if i < len(events) else 0)
                    ws4.cell(row=row, column=3, value=created[i] if i < len(created) else 0)
                    ws4.cell(row=row, column=4, value=closed[i] if i < len(closed) else 0)
                    ws4.cell(row=row, column=5, value=critical[i] if i < len(critical) else 0)
                    ws4.cell(row=row, column=6, value=important[i] if i < len(important) else 0)

                if dates:
                    style_data(ws4, 4, 4 + len(dates) - 1, 6)
                for col in range(1, 7):
                    ws4.column_dimensions[get_column_letter(col)].width = [15, 15, 15, 15, 15, 15][col-1]

            wb.save(filepath)
            return filepath

        except Exception as e:
            self.logger.error("Failed to generate Excel report", error=str(e))
            fallback_path = os.path.join(
                settings.REPORT_OUTPUT_DIR,
                f"合规日报_{report_date.strftime('%Y%m%d')}_excel_data.json"
            )
            with open(fallback_path, "w", encoding="utf-8") as f:
                json.dump({"error": str(e), "date": report_date.isoformat()}, f, ensure_ascii=False, indent=2)
            return fallback_path

    @staticmethod
    def _event_type_label(event_type_value: str) -> str:
        try:
            et = EventType(event_type_value)
            cn_map = {
                EventType.DATA_LEAK: "数据泄露",
                EventType.UNAUTHORIZED_ACCESS: "未授权访问",
                EventType.FRAUD: "财务欺诈",
                EventType.CONFLICT_OF_INTEREST: "利益冲突",
                EventType.HARASSMENT: "职场骚扰",
                EventType.DISCRIMINATION: "歧视行为",
                EventType.THEFT: "盗窃行为",
                EventType.POLICY_VIOLATION: "违反制度",
                EventType.SUSPICIOUS_COMMUNICATION: "可疑通讯",
                EventType.ABNORMAL_BEHAVIOR: "异常行为",
                EventType.OTHER: "其他",
            }
            return cn_map.get(et, event_type_value)
        except Exception:
            return event_type_value
